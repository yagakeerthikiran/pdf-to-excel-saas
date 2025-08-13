import os
import glob
import pytest
import openpyxl
from backend.conversion_service import convert_pdf_to_excel
from pypdf.errors import FileNotDecryptedError
from pdfplumber.utils.exceptions import PdfminerException

# --- Test Configuration ---
PDF_SAMPLES_DIR = "/tmp/pdf-test-files"
TEST_OUTPUT_DIR = "tests/output"

if os.path.isdir(PDF_SAMPLES_DIR):
    pdf_files = glob.glob(os.path.join(PDF_SAMPLES_DIR, "*.pdf"))
else:
    pdf_files = []

@pytest.fixture(scope="session", autouse=True)
def setup_teardown_session():
    os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
    yield

@pytest.mark.parametrize("pdf_path", pdf_files)
def test_vision_assisted_conversion_pipeline(pdf_path):
    """
    Tests the vision-assisted conversion pipeline on a given PDF file.
    """
    assert os.path.exists(pdf_path), f"Sample PDF not found: {pdf_path}"

    base_filename = os.path.splitext(os.path.basename(pdf_path))[0]
    excel_path = os.path.join(TEST_OUTPUT_DIR, f"{base_filename}_output.xlsx")

    try:
        convert_pdf_to_excel(pdf_path, excel_path)
    except (FileNotDecryptedError, PdfminerException) as e:
        pytest.skip(f"Skipping un-processable PDF file: {pdf_path}. Reason: {type(e).__name__}")
    except Exception as e:
        pytest.fail(f"convert_pdf_to_excel raised an unexpected exception for {pdf_path}: {e}")

    # --- Assertions ---
    assert os.path.exists(excel_path), f"Output Excel file was not created for {pdf_path}"
    assert os.path.getsize(excel_path) > 0, f"The output Excel file for {pdf_path} is empty."

    # --- Specific Assertions for the NAB Statement ---
    if "NAB" in base_filename:
        try:
            workbook = openpyxl.load_workbook(excel_path)
            # Assuming the main table is the first one found
            table_sheet = workbook["Table_1_Page_1"]

            # Find the header row
            header_row_idx = -1
            details_col_idx = -1
            debits_col_idx = -1

            for i, row in enumerate(table_sheet.iter_rows(values_only=True), start=1):
                if "Transaction Details" in row and "Debits" in row:
                    header_row_idx = i
                    # Find the column indices for the headers
                    details_col_idx = row.index("Transaction Details")
                    debits_col_idx = row.index("Debits")
                    break

            assert header_row_idx != -1, "Could not find the header row in the NAB statement."
            assert details_col_idx != debits_col_idx, "Transaction Details and Debits are in the same column."

            # Find a specific transaction and verify alignment
            found_transaction = False
            for row in table_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                # Using a known transaction from a sample file
                if row[details_col_idx] and "TRANSFER" in row[details_col_idx] and "VODAFONE" in row[details_col_idx]:
                    # This is the row we want to check
                    # The value in the "Debits" column for this row should be a number
                    debit_value = row[debits_col_idx]
                    assert isinstance(debit_value, (int, float)) or debit_value.replace('$', '').replace(',', '').strip().isdigit()
                    found_transaction = True
                    break

            assert found_transaction, "Could not find the specific VODAFONE transaction to verify alignment."

        except Exception as e:
            pytest.fail(f"Failed to inspect the output Excel file for {pdf_path}: {e}")
