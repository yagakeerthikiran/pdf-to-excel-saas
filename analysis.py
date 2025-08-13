import openpyxl
from backend.conversion_service import convert_pdf_to_excel
import os

# --- Configuration ---
PDF_SAMPLES_DIR = "/tmp/pdf-test-files"
ANALYSIS_OUTPUT_DIR = "/tmp/analysis_output"
os.makedirs(ANALYSIS_OUTPUT_DIR, exist_ok=True)

SOURCE_PDF = os.path.join(PDF_SAMPLES_DIR, "Kiran_SBI_Statement.pdf")
TARGET_EXCEL = os.path.join(PDF_SAMPLES_DIR, "Kiran_SBI_Statement.xlsx")
MY_OUTPUT_EXCEL = os.path.join(ANALYSIS_OUTPUT_DIR, "my_output.xlsx")

# --- 1. Generate our own version ---
print("--- Generating our version of the Excel file... ---")
try:
    convert_pdf_to_excel(SOURCE_PDF, MY_OUTPUT_EXCEL)
    print(f"Successfully generated our output at: {MY_OUTPUT_EXCEL}")
except Exception as e:
    print(f"Error generating our output: {e}")
    exit()

# --- 2. Analyze both Excel files ---
def analyze_excel(file_path):
    print(f"\\n--- Analyzing: {os.path.basename(file_path)} ---")
    try:
        workbook = openpyxl.load_workbook(file_path)
        print(f"Sheet Names: {workbook.sheetnames}")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"  Sheet '{sheet_name}':")
            print(f"    - Dimensions: {sheet.dimensions}")
            if sheet.merged_cells:
                print(f"    - Merged Cells: {sheet.merged_cells}")
            else:
                print("    - No merged cells.")

            # Print first 5 rows for a sample
            print("    - Sample Data (first 5 rows):")
            for row in sheet.iter_rows(min_row=1, max_row=5):
                row_data = [cell.value for cell in row if cell.value is not None]
                if row_data:
                    print(f"      {row_data}")

    except Exception as e:
        print(f"Could not analyze file {file_path}. Error: {e}")

analyze_excel(TARGET_EXCEL)
analyze_excel(MY_OUTPUT_EXCEL)

# --- 3. Manual Comparison Summary (based on expected output) ---
print("\\n--- Comparison Summary ---")
print("Key expected differences to investigate:")
print("1. Sheet Structure: Does the target file use one sheet or multiple? My current logic creates multiple sheets ('All_Text', 'Table_1', etc.). The competitor might be combining everything smartly into one.")
print("2. Merged Cells: The target file likely uses merged cells for headers and titles to mimic the PDF layout. My current logic does not merge any cells.")
print("3. Data Placement: The target file probably places text and tables on the same sheet in a layout that mirrors the PDF. My logic completely separates them.")
print("4. Empty Rows/Columns: The target might be inserting empty rows and columns to create visual spacing, just like in the PDF.")
print("--------------------------")
print("\\nBased on this analysis, I will formulate a new plan to improve the conversion service.")
