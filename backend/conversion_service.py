import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter
import structlog
from typing import List, Dict, Any, Optional
import backend.ocr_service as ocr_service

logger = structlog.get_logger(__name__)

Element = Dict[str, Any]

def parse_textract_blocks(textract_json: Optional[dict], page_width: float, page_height: float) -> List[Element]:
    """
    Parses the Textract JSON response and converts it into our Element format.

    Args:
        textract_json: Parsed JSON returned by AWS Textract. Some callers may
            provide ``None`` if OCR failed or was skipped.
        page_width: Width of the PDF page in points.
        page_height: Height of the PDF page in points.

    Returns:
        A list of element dictionaries describing the words detected on the
        page. Returns an empty list when ``textract_json`` is ``None``.
    """
    if not textract_json:
        return []

    elements = []
    for block in textract_json.get("Blocks", []):
        if block["BlockType"] == "WORD":
            text = block.get("Text", "")
            geom = block.get("Geometry", {})
            bbox = geom.get("BoundingBox", {})

            # Convert Textract's proportional coordinates to absolute pdfplumber-style coordinates
            x0 = bbox.get("Left", 0) * page_width
            top = bbox.get("Top", 0) * page_height
            x1 = x0 + (bbox.get("Width", 0) * page_width)
            bottom = top + (bbox.get("Height", 0) * page_height)

            elements.append({
                "text": text,
                "x0": x0, "top": top, "x1": x1, "bottom": bottom,
                "height": bottom - top
            })
    return elements

def build_grid_from_elements(elements: List[Element]) -> List[List[str]]:
    """
    Analyzes the layout of OCR-extracted word elements and arranges them into a grid.
    """
    if not elements:
        return []

    # A rough estimate for average character width, used for tolerances
    avg_char_width = 0
    if elements:
        # Filter out elements with no text to avoid division by zero
        valid_elements = [el for el in elements if el['text']]
        if valid_elements:
            avg_char_width = sum((el['x1'] - el['x0']) / len(el['text']) for el in valid_elements) / len(valid_elements)

    # 1. Group elements into lines based on vertical position
    lines_of_elements = []
    if elements:
        elements.sort(key=lambda el: (el['top'], el['x0']))
        current_line = [elements[0]]
        for el in elements[1:]:
            prev_el = current_line[-1]
            line_break_threshold = prev_el['height'] * 0.7
            if el['top'] > (prev_el['top'] + line_break_threshold):
                lines_of_elements.append(current_line)
                current_line = [el]
            else:
                current_line.append(el)
        lines_of_elements.append(current_line)

    # 2. Pre-process lines to merge adjacent words
    merged_lines = []
    for line in lines_of_elements:
        if not line: continue
        merged_line = []
        current_word = dict(line[0]) # Make a copy
        for next_word in line[1:]:
            # If next word is close enough, merge it
            gap = next_word['x0'] - current_word['x1']
            if gap < avg_char_width * 0.8:
                current_word['text'] += ' ' + next_word['text']
                current_word['x1'] = next_word['x1']
            else:
                merged_line.append(current_word)
                current_word = dict(next_word)
        merged_line.append(current_word)
        merged_lines.append(merged_line)

    # 3. Determine column boundaries from the merged elements
    column_starts = []
    all_merged_elements = [el for line in merged_lines for el in line]
    if all_merged_elements:
        sorted_x0s = sorted(list(set(el['x0'] for el in all_merged_elements)))
        if sorted_x0s:
            column_starts.append(sorted_x0s[0])
            for x0 in sorted_x0s[1:]:
                # If a new x0 is reasonably far from the last column start, it's a new column
                if x0 > (column_starts[-1] + avg_char_width):
                    column_starts.append(x0)

    # 4. Build the grid using the merged lines and detected columns
    grid = []
    for line in merged_lines:
        row = [""] * len(column_starts)
        for el in line:
            # Find the index of the column this element belongs to
            target_col_index = -1
            # Find the column start that is closest to the element's start
            min_dist = float('inf')
            for i, col_x0 in enumerate(column_starts):
                dist = abs(el['x0'] - col_x0)
                # Prefer columns that are to the left of the element or not too far right
                if dist < min_dist and el['x0'] + avg_char_width > col_x0:
                    min_dist = dist
                    target_col_index = i

            if target_col_index != -1:
                row[target_col_index] = el['text']
        grid.append(row)

    return grid

def map_elements_to_sheet(sheet, elements: List[Element]):
    """
    This function is now much simpler. It just places pre-structured table data.
    The complex layout logic is now handled by the extraction strategy.
    """
    for r, row_data in enumerate(elements, start=1):
        for c, cell_data in enumerate(row_data, start=1):
            cleaned_data = str(cell_data).replace('\\n', ' ').strip() if cell_data else ""
            sheet.cell(row=r, column=c).value = cleaned_data

    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].auto_size = True

def convert_pdf_to_excel(pdf_path: str, excel_path: str, bucket_name: str, object_key: str):
    """
    The final, 3-tiered hybrid conversion engine.
    """
    logger.info("Starting 3-tier hybrid conversion", pdf_path=pdf_path)

    try:
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        with pdfplumber.open(pdf_path) as pdf:
            # --- Tier 0: Full text extraction as a fallback sheet ---
            text_sheet = workbook.create_sheet(title="All_Text_Fallback")
            full_text = "".join([page.extract_text(x_tolerance=2) or "" for page in pdf.pages])
            text_sheet.cell(row=1, column=1).value = full_text

            # --- Tier 3 Setup: Pre-run OCR if necessary ---
            is_image_pdf = not any(p.extract_text(x_tolerance=2).strip() for p in pdf.pages)
            textract_json = None
            if "NAB" in pdf_path or is_image_pdf:
                logger.info("Document may require OCR; running AWS Textract.", pdf_path=pdf_path)
                try:
                    textract_json = ocr_service.extract_data_with_textract(bucket_name, object_key)
                except Exception as e:
                    logger.error("AWS Textract call failed", error=str(e))
                    workbook.create_sheet(title="OCR_Error").cell(row=1, column=1).value = f"Textract failed: {e}"

            blocks_by_page = {}
            if textract_json:
                for block in textract_json.get("Blocks", []):
                    page_num = block.get("Page", 1)
                    blocks_by_page.setdefault(page_num, []).append(block)

            # --- Main Page Loop ---
            table_count = 0
            for i, page in enumerate(pdf.pages):
                page_num = i + 1
                logger.info(f"Processing Page {page_num}...")

                # Heuristic to decide if we should use OCR data for this page
                use_ocr = (page_num in blocks_by_page) and ("NAB" in pdf_path or not page.extract_text(x_tolerance=2).strip())

                if use_ocr:
                    logger.info(f"Using OCR data for page {page_num}.")
                    page_blocks = blocks_by_page[page_num]
                    ocr_sheet = workbook.create_sheet(title=f"OCR_Page_{page_num}")

                    elements = parse_textract_blocks({"Blocks": page_blocks}, page.width, page.height)
                    grid = build_grid_from_elements(elements)
                    map_elements_to_sheet(ocr_sheet, grid)
                    continue # Skip to next page

                # --- Tiers 1 & 2: pdfplumber table extraction ---
                logger.info(f"Using pdfplumber for page {page_num}.")
                line_tables = page.extract_tables({"vertical_strategy": "lines", "horizontal_strategy": "lines"})
                if not line_tables:
                    logger.warn(f"No tables found with 'lines' strategy on page {page_num}. Falling back to 'text'.")
                    tables_to_write = page.extract_tables({"vertical_strategy": "text", "horizontal_strategy": "text"})
                else:
                    tables_to_write = line_tables

                if not tables_to_write:
                    logger.warn(f"No tables found by pdfplumber on page {page_num}.")

                for table in tables_to_write:
                    table_count += 1
                    table_sheet = workbook.create_sheet(title=f"Table_pdfplumber_{table_count}_P{page_num}")
                    map_elements_to_sheet(table_sheet, table)

        workbook.save(excel_path)
        logger.info("Successfully completed hybrid conversion.", excel_path=excel_path)

    except Exception as e:
        logger.error("Failed during hybrid conversion", error=str(e), exc_info=True)
        raise
